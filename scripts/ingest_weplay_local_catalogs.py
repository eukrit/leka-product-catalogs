"""Mine the local Google-Drive Weplay catalog folder for additional EN
content covering older / discontinued SKUs that the live + cached + 2025
flipbook scrapes missed.

Source folder: C:\\Users\\Eukrit\\My Drive\\Catalogs GO\\WePlay Catalogs\\

Sources used (text-extractable only — image-only PDFs need Vision OCR
and are skipped here):
  - *Pricelist*.xlsx                    : Model + Desc + List/Discount/EXW
  - *Quotation*.pdf                     : SKU + name + qty + price (4p each)
  - Weplay Catalogue 2021-2022.pdf      : text layer present (~1.2k chars/p)
  - Weplay Catalogue 2023-2024.pdf      : text layer present (~0.9k chars/p)

Skipped (image-only, would need Vision):
  - 2025 catalog (153MB, 95p, 7 chars/p)
  - 2020-2021 catalog (75MB, 83p, 135 chars/p — mostly imgs)
  - 2022-2023 catalog (183MB, 91p, 0 chars/p)
  - New Products 2021-2023.pdf (12MB, 61p, 18 chars/p)

Idempotent. Only writes when the doc has no live/cached/flipbook source
yet. Provenance saved as `source_url_local` = "<filename>:<page>" or
"<filename>:row<N>".

Usage:
    py scripts/ingest_weplay_local_catalogs.py --dry-run
    py scripts/ingest_weplay_local_catalogs.py --apply
    py scripts/ingest_weplay_local_catalogs.py --apply --dump=/tmp/local.json
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from pathlib import Path

import openpyxl
import pdfplumber
from google.cloud import firestore

_FALLBACK_ADC = (
    r"C:\Users\Eukrit\AppData\Roaming\gcloud\legacy_credentials"
    r"\codex-chatgpt@ai-agents-go.iam.gserviceaccount.com\adc.json"
)
if "GOOGLE_APPLICATION_CREDENTIALS" not in os.environ or not os.path.exists(
    os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "")
):
    if os.path.exists(_FALLBACK_ADC):
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = _FALLBACK_ADC
os.environ.setdefault("GOOGLE_CLOUD_PROJECT", "ai-agents-go")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger("ingest_weplay_local_catalogs")

PROJECT = "ai-agents-go"
DB = "vendors"
SLUG = "weplay"
FOLDER = Path(r"C:\Users\Eukrit\My Drive\Catalogs GO\WePlay Catalogs")
SKU_TOKEN_RE = re.compile(r"([A-Z]{2}[0-9]{4,})")
# Stricter version that requires the SKU to be the WHOLE model code (allows
# dot/dash suffix). Used for Excel "Model" cells where the entire cell is the SKU.
WHOLE_SKU_RE = re.compile(r"^([A-Z]{2}[0-9]{4,}(?:[.-][A-Z0-9]+)?)$")


# -------- Excel pricelist parsing -------------------------------------------

def parse_excel_pricelist(path: Path) -> list[dict]:
    """Return [{sku, name_en, list_price, exw_price, source}]."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find header row — first row with both "Model" and "Desc"
        header_idx = None
        col_map: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
            cells = [str(c).strip() if c is not None else "" for c in row]
            joined = " | ".join(c.lower() for c in cells)
            if "model" in joined and "desc" in joined:
                header_idx = r_idx
                for c_idx, c in enumerate(cells):
                    cl = c.lower().strip()
                    if cl in ("model", "sku", "item code"):
                        col_map["sku"] = c_idx
                    elif cl in ("desc", "description", "product"):
                        col_map["desc"] = c_idx
                    elif cl in ("list", "list price", "msrp"):
                        col_map["list"] = c_idx
                    elif cl in ("exw", "exw price", "fob"):
                        col_map["exw"] = c_idx
                break
        if header_idx is None or "sku" not in col_map:
            log.info("  '%s/%s': no header row found, skipping", path.name, sheet_name)
            continue
        # Iterate data rows
        for r_idx, row in enumerate(ws.iter_rows(min_row=header_idx + 2, values_only=True)):
            sku_cell = row[col_map["sku"]] if col_map["sku"] < len(row) else None
            if sku_cell is None:
                continue
            sku_raw = str(sku_cell).strip().upper()
            if not WHOLE_SKU_RE.match(sku_raw):
                # Sometimes the model cell has prefix like "6800KM1003"; try inner match
                m = SKU_TOKEN_RE.search(sku_raw)
                if not m:
                    continue
                sku_for_token = m.group(1)
            else:
                sku_for_token = WHOLE_SKU_RE.match(sku_raw).group(1)
            token_m = SKU_TOKEN_RE.search(sku_for_token)
            token = token_m.group(1) if token_m else sku_for_token
            desc_cell = row[col_map.get("desc", -1)] if col_map.get("desc", -1) < len(row) else None
            desc = str(desc_cell).strip() if desc_cell else ""
            list_p = row[col_map.get("list", -1)] if col_map.get("list", -1) >= 0 and col_map["list"] < len(row) else None
            exw_p = row[col_map.get("exw", -1)] if col_map.get("exw", -1) >= 0 and col_map["exw"] < len(row) else None
            out.append({
                "sku": sku_raw,
                "sku_token": token,
                "name_en": desc[:120] if desc else "",
                "description_en": "",  # pricelists usually only have name in Desc
                "list_price": float(list_p) if isinstance(list_p, (int, float)) else None,
                "exw_price": float(exw_p) if isinstance(exw_p, (int, float)) else None,
                "source": f"{path.name}:{sheet_name}:row{header_idx + 2 + r_idx}",
            })
    log.info("  parsed %d rows from %s", len(out), path.name)
    return out


# -------- PDF parsing -------------------------------------------------------

def parse_pdf(path: Path, max_pages: int | None = None) -> list[dict]:
    """Extract SKU + nearby text from a text-layer PDF."""
    out: list[dict] = []
    with pdfplumber.open(path) as pdf:
        n = min(len(pdf.pages), max_pages) if max_pages else len(pdf.pages)
        for pn in range(n):
            text = pdf.pages[pn].extract_text() or ""
            if not text.strip():
                continue
            # For each SKU mention, capture name candidates from surrounding
            # text. Use heuristic: SKU is typically followed by a product name
            # on the same line (e.g. "KM1003 Pile Balance Up").
            for m in SKU_TOKEN_RE.finditer(text):
                token = m.group(1)
                start = m.start()
                # Capture context: from start of line containing match to end of
                # next 200 chars or end of paragraph.
                line_start = text.rfind("\n", 0, start) + 1
                line_end = text.find("\n", start + 200)
                if line_end < 0:
                    line_end = min(len(text), start + 400)
                context = text[line_start:line_end].strip()
                # Try to extract the name: text after the SKU on the same line
                after_sku = text[m.end(): m.end() + 120]
                # First 40 non-numeric chars is usually the name
                name_match = re.match(r"\s*[:\-]?\s*([A-Z][\w\s,&'/().\-]+?)(?:\s{2,}|\d{2,}|$)", after_sku)
                name = name_match.group(1).strip() if name_match else ""
                # Truncate name to reasonable length
                name = name[:80].strip()
                if not name or len(name) < 4:
                    continue
                out.append({
                    "sku": token,
                    "sku_token": token,
                    "name_en": name,
                    "description_en": context[:300],
                    "source": f"{path.name}:p{pn + 1}",
                })
    log.info("  parsed %d entries from %s", len(out), path.name)
    return out


# -------- Aggregation + writeback -------------------------------------------

def aggregate(records: list[dict]) -> dict[str, dict]:
    """Dedupe by token; keep richest entry."""
    by_token: dict[str, dict] = {}
    for r in records:
        token = r["sku_token"]
        existing = by_token.get(token)
        if not existing:
            by_token[token] = dict(r, sources=[r["source"]])
            continue
        # Merge: keep longer description, longer name, append source
        existing["sources"].append(r["source"])
        if len(r.get("description_en") or "") > len(existing.get("description_en") or ""):
            existing["description_en"] = r["description_en"]
        if len(r.get("name_en") or "") > len(existing.get("name_en") or ""):
            existing["name_en"] = r["name_en"]
        if r.get("list_price") and not existing.get("list_price"):
            existing["list_price"] = r["list_price"]
        if r.get("exw_price") and not existing.get("exw_price"):
            existing["exw_price"] = r["exw_price"]
    return by_token


def writeback(by_token: dict[str, dict], dry_run: bool) -> dict:
    db = firestore.Client(project=PROJECT, database=DB)
    coll = db.collection("vendors").document(SLUG).collection("products")
    token_to_docs: dict[str, list[firestore.DocumentSnapshot]] = {}
    for snap in coll.stream():
        d = snap.to_dict() or {}
        sku = (d.get("item_code") or "").upper()
        m = SKU_TOKEN_RE.search(sku)
        if not m:
            continue
        token_to_docs.setdefault(m.group(1), []).append(snap)

    counters = {
        "local_skus": len(by_token),
        "matched_to_doc": 0,
        "skipped_already_has_better_source": 0,
        "writes": 0,
        "no_doc_match": 0,
    }
    sample_writes, sample_no_match = [], []
    batch = db.batch()
    batch_n = 0
    BATCH = 200

    for token, det in by_token.items():
        targets = token_to_docs.get(token, [])
        if not targets:
            counters["no_doc_match"] += 1
            if len(sample_no_match) < 8:
                sample_no_match.append(f"{det['sku']}: {det.get('name_en')!r}")
            continue
        for snap in targets:
            counters["matched_to_doc"] += 1
            d = snap.to_dict() or {}
            # Don't clobber any source already set (live, cached, flipbook).
            if d.get("source_url_en") or d.get("source_url_cached") or d.get("source_url_flipbook"):
                counters["skipped_already_has_better_source"] += 1
                continue

            orig_name = d.get("product_name") or ""
            is_chinese_orig = bool(re.search(r"[一-鿿]", orig_name))
            payload = {
                "name": (det.get("name_en") or "").strip() or d.get("name"),
                "description": (det.get("description_en") or "").strip() or d.get("description"),
                "source_url_local": ", ".join(det.get("sources", [])[:3]),
                "description_orig": d.get("description"),
            }
            if is_chinese_orig:
                payload["name_zh"] = orig_name
            # If pricelist had a list/EXW price, include it for reference
            if det.get("list_price") or det.get("exw_price"):
                pricing = dict(d.get("pricing") or {})
                if det.get("list_price"):
                    pricing["list_2021"] = det["list_price"]
                if det.get("exw_price"):
                    pricing["exw_2021"] = det["exw_price"]
                payload["pricing"] = pricing

            counters["writes"] += 1
            if len(sample_writes) < 8:
                sample_writes.append(
                    f"{snap.id} sku={det['sku']} name={(payload.get('name') or '')[:50]!r} src={det['sources'][0]}"
                )
            if not dry_run:
                batch.set(snap.reference, payload, merge=True)
                batch_n += 1
                if batch_n >= BATCH:
                    batch.commit()
                    log.info("committed batch (%d docs)", batch_n)
                    batch = db.batch()
                    batch_n = 0

    if not dry_run and batch_n:
        batch.commit()
        log.info("committed final batch (%d docs)", batch_n)

    log.info("=== summary ===")
    for k, v in counters.items():
        log.info("  %s: %d", k, v)
    if sample_writes:
        log.info("sample writes:")
        for s in sample_writes:
            log.info("  %s", s)
    if sample_no_match:
        log.info("sample no-doc-match: %s", sample_no_match)
    return counters


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    mode = ap.add_mutually_exclusive_group(required=True)
    mode.add_argument("--apply", action="store_true")
    mode.add_argument("--dry-run", action="store_true")
    ap.add_argument("--dump", type=str, default=None)
    args = ap.parse_args()
    write = bool(args.apply)
    log.info("=== ingest_weplay_local_catalogs mode=%s ===", "WRITE" if write else "DRY-RUN")

    if not FOLDER.exists():
        log.error("folder not found: %s", FOLDER)
        return 2

    records: list[dict] = []

    log.info("\n--- Excel pricelists ---")
    for x in sorted(FOLDER.glob("*Pricelist*.xlsx")):
        records.extend(parse_excel_pricelist(x))

    log.info("\n--- Text-layer PDFs ---")
    for pdf_name in (
        "Weplay Catalogue 2021-2022.pdf",
        "Weplay Catalogue 2023-2024.pdf",
    ):
        path = FOLDER / pdf_name
        if path.exists():
            records.extend(parse_pdf(path))

    log.info("\n--- Quotation PDFs ---")
    for q in sorted(FOLDER.glob("*Quotation*.pdf")):
        records.extend(parse_pdf(q))

    by_token = aggregate(records)
    log.info("aggregated: %d unique SKU tokens (from %d raw records)",
             len(by_token), len(records))

    if args.dump:
        with open(args.dump, "w", encoding="utf-8") as f:
            json.dump(by_token, f, indent=2, ensure_ascii=False, default=str)
        log.info("dumped %d to %s", len(by_token), args.dump)

    writeback(by_token, dry_run=not write)
    return 0


if __name__ == "__main__":
    sys.exit(main())
