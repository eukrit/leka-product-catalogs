"""Phase B1: parse DesignPark's 2024 USD pricelist + 2023 theme manifest
into `vendors/designpark/products/<handle>` docs.

Sources (local Google Drive):
  A. 2024-03-27 D'Park Price List (USD-2024).xlsx
     - 12 sheets, ~300 SKUs, USD FOB Busan unit prices.
     - Each sheet = one product family (Slides & Tubes, Fitness *, Speed
       Racers, Play (Dry/GRC/Aquatic), Modern Igloo).
     - Header row layout: No | IMAGE | MODEL NO | DESCRIPTION |
       UNIT PRICE (EX_WORK | FOB Busan) | REMARKS
  B. 2024-03-18 D'Park 2D CAD & Images / 2023 Theme dry&waterplay list ... .xlsx
     - 1 sheet, ~33 themes.
     - Layout: NO | CATEGORY | THEME | PRODUCT NAME (EN) | IMAGE | SIZE |
       MAIN MATERIAL | REMARK + many color/option columns.
     - No per-theme price; treated as bundle catalog (status="draft_no_price").

Pricing model (per plan §7, decision #4 — USD + THB + EUR Vinci-style):
  - cost_engine origin = "japan_korea" (LCL Busan → Bangkok).
  - duty_rate = DUTY_RATE_NON_CHINA (0.10) — DesignPark is Korea, not China.
  - GROSS_MARGIN = 0.35 (matches Vinci/landed_pricing default; brand override
    possible later via root-doc field if user revises).
  - When SKU has no dimension data (every line in this pricelist — no
    L/W/H column in the workbook), price_row() falls back to the flat
    UNMATCHED_LANDED_UPLIFT = 1.35 path inside landed_pricing.py.
    We feed USD via the FX layer so EUR FOB = usd * USD/EUR.

Idempotent. Merge-writes only — does not delete existing fields.
Provenance fields written:
  source_pricelist:  "<filename>:<sheet>:row<N>"
  category:          sheet name (e.g. "Slides & Tubes")
  pricing.fob_usd:   from pricelist
  pricing.landed_thb / retail_thb / retail_usd / retail_eur: computed
  pricing.formula_version: "designpark-v1-2026-05-15"

Usage:
    py scripts/ingest_designpark_pricelist.py --dry-run
    py scripts/ingest_designpark_pricelist.py --apply
    py scripts/ingest_designpark_pricelist.py --apply --skip-themes
    py scripts/ingest_designpark_pricelist.py --apply --limit=10
    py scripts/ingest_designpark_pricelist.py --dry-run --dump-csv=/tmp/dp.csv
"""
from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
from pathlib import Path

# ADC bootstrap — mirrors bootstrap_designpark.py.
_LOCAL_SA_CANDIDATES = [
    r"C:\Users\Eukrit\OneDrive\Documents\Claude Code\Credentials Claude Code\ai-agents-go-9b4219be8c01.json",
    r"C:\Users\eukri\OneDrive\Documents\Claude Code\Credentials Claude Code\ai-agents-go-9b4219be8c01.json",
]
if "GOOGLE_APPLICATION_CREDENTIALS" not in os.environ:
    for cand in _LOCAL_SA_CANDIDATES:
        if os.path.exists(cand):
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = cand
            break
os.environ.setdefault("GOOGLE_CLOUD_PROJECT", "ai-agents-go")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from google.cloud import firestore  # noqa: E402

from shared.landed_pricing import (  # noqa: E402
    DEFAULT_PACKING_FACTOR,
    GROSS_MARGIN,
    THAI_VAT_RATE,
    TH_CUSTOMER_VAT_RATE,
    UNMATCHED_LANDED_UPLIFT,
    DUTY_RATE_NON_CHINA,
    SG_CUSTOMER_GST_RATE,
    SG_NUBO_GST_REGISTERED,
    calibrate_baltic_rate,
    get_fx_rates,
)
from shared.pricing_config import get_pricing_config  # noqa: E402
import cost_engine  # noqa: E402 — resolved via shared.landed_pricing's sys.path nudge

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger("ingest_designpark_pricelist")

PROJECT = "ai-agents-go"
VENDORS_DB = "vendors"
SLUG = "designpark"
FORMULA_VERSION = "designpark-v1-2026-05-15"

# --- Source files (cross-machine; supports both Eukrit/eukri profiles) ------
DRIVE_PARTNER = Path(r"C:\Users\Eukrit\My Drive\Partners Playground\DesignPark")
PRICELIST_USD = DRIVE_PARTNER / "2024-03-27 Design Park Pricelist D'Park Price List (USD-2024).xlsx"
THEME_MANIFEST = (
    DRIVE_PARTNER
    / "2024-03-18 D'Park 2D CAD & Images"
    / "2023 Theme dry&waterplay list in 2022_Designpark_240313.xlsx"
)

# --- Parsing ---------------------------------------------------------------
HEADER_ALIASES = {
    "no": "row_no",
    "no.": "row_no",
    "image": "image_cell",
    "model no": "item_code",
    "model no.": "item_code",
    "model number": "item_code",
    "description": "description",
    "remarks": "remarks",
    "remark": "remarks",
}

# Any header containing "unit price" → unit_price (currency derived from text).
UNIT_PRICE_RE = re.compile(r"unit\s*price", re.IGNORECASE)
SLUG_BAD_CHARS = re.compile(r"[^a-z0-9]+")


def slugify(text: str) -> str:
    s = SLUG_BAD_CHARS.sub("-", text.lower()).strip("-")
    return s or "unknown"


def to_handle(item_code: str, name: str) -> str:
    """Stable handle. Prefer item_code (already SKU-shaped) when present."""
    base = item_code or name
    return f"designpark-{slugify(base)}"


def parse_sheet(sheet_name: str, ws) -> list[dict]:
    """Find the header row in the first 12 rows, then iterate data rows."""
    header_idx: int | None = None
    col_map: dict[str, int] = {}
    for r_idx in range(1, min(13, ws.max_row + 1)):
        row = [c.value for c in ws[r_idx]]
        cells_lc = [(str(c).strip().lower() if c is not None else "") for c in row]
        if "model no" in cells_lc or "model no." in cells_lc:
            header_idx = r_idx
            for col_idx, cell in enumerate(cells_lc):
                if not cell:
                    continue
                if cell in HEADER_ALIASES:
                    col_map[HEADER_ALIASES[cell]] = col_idx
                elif UNIT_PRICE_RE.search(cell):
                    col_map["unit_price"] = col_idx
                    col_map["unit_price_label"] = col_idx  # remember text too
            break
    if header_idx is None or "item_code" not in col_map or "unit_price" not in col_map:
        # Fallback path for category-only sheets like "Modern Igloo" that lack
        # a MODEL NO column. We scan for a header with Category + Description
        # + Unit Price (no Model No) and synthesize SKUs from the sheet name.
        if "item_code" not in (col_map or {}) and header_idx is None:
            for r_idx in range(1, min(13, ws.max_row + 1)):
                row = [c.value for c in ws[r_idx]]
                cells_lc = [(str(c).strip().lower() if c is not None else "") for c in row]
                has_desc = "description" in cells_lc
                has_cat = "category" in cells_lc
                has_price = any(UNIT_PRICE_RE.search(c) for c in cells_lc if c)
                if has_desc and has_price and (has_cat or sheet_name.lower() in ("modern igloo",)):
                    header_idx = r_idx
                    col_map = {}
                    for col_idx, cell in enumerate(cells_lc):
                        if not cell:
                            continue
                        if cell == "description":
                            col_map["description"] = col_idx
                        elif cell == "category":
                            col_map["category"] = col_idx
                        elif UNIT_PRICE_RE.search(cell):
                            col_map["unit_price"] = col_idx
                    col_map["item_code"] = -1  # synthetic
                    break
        if header_idx is None or "unit_price" not in col_map:
            log.warning("[%s] no parseable header — skipping (max_row=%s)", sheet_name, ws.max_row)
            return []

    rows: list[dict] = []
    synthetic_sku = col_map.get("item_code") == -1
    for r_idx in range(header_idx + 1, ws.max_row + 1):
        row = [c.value for c in ws[r_idx]]
        if not row:
            continue

        desc = row[col_map["description"]] if "description" in col_map else ""
        desc_s = str(desc).strip() if desc is not None else ""

        if synthetic_sku:
            # No MODEL NO column — derive item_code from sheet name + description.
            if not desc_s or desc_s.lower().startswith(("1)", "2)", "3)", "4)", "5)", "this does not")):
                # Skip note/footer rows.
                continue
            item_code_s = "DP-" + slugify(sheet_name).upper() + "-" + slugify(desc_s).upper()[:32]
        else:
            item_code = row[col_map["item_code"]]
            if item_code is None or str(item_code).strip() == "":
                continue
            item_code_s = str(item_code).strip()

        price = row[col_map["unit_price"]]
        if isinstance(price, str):
            price_clean = re.sub(r"[^\d.]", "", price)
            try:
                fob_usd = float(price_clean) if price_clean else None
            except ValueError:
                fob_usd = None
        elif isinstance(price, (int, float)):
            fob_usd = float(price)
        else:
            fob_usd = None

        if fob_usd is None or fob_usd <= 0:
            log.debug("[%s row %d] no price for %s", sheet_name, r_idx, item_code_s)
            # Still include — price-less components are valid catalog rows.
            fob_usd = 0.0

        remarks_v = row[col_map["remarks"]] if "remarks" in col_map else ""
        remarks_s = str(remarks_v).strip() if remarks_v else ""

        rows.append({
            "category": sheet_name,
            "item_code": item_code_s,
            "name": desc_s or item_code_s,
            "description": desc_s,
            "fob_usd": fob_usd,
            "remarks": remarks_s,
            "source_pricelist": f"{PRICELIST_USD.name}:{sheet_name}:row{r_idx}",
        })
    log.info("[%s] parsed %d rows", sheet_name, len(rows))
    return rows


def parse_theme_manifest() -> list[dict]:
    if not THEME_MANIFEST.exists():
        log.warning("theme manifest not found: %s", THEME_MANIFEST)
        return []
    wb = openpyxl.load_workbook(THEME_MANIFEST, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Header row is row 2: NO | CATEGORY | THEME | PRODUCT NAME (EN) | IMAGE | SIZE | MAIN MATERIAL | REMARK
    header_idx = 2
    rows: list[dict] = []
    for r_idx in range(header_idx + 1, ws.max_row + 1):
        row = [c.value for c in ws[r_idx]]
        if len(row) < 8:
            continue
        no, category, theme, name_en, _img, size, material, remark = row[:8]
        if not name_en or str(name_en).strip() == "":
            continue
        name_s = str(name_en).strip()
        cat_s = str(category).strip() if category else "Themes"
        theme_s = str(theme).strip() if theme else ""
        size_s = str(size).strip() if size else ""
        mat_s = str(material).strip() if material else ""
        # Theme manifest has no item_code column. Synthesize: DP-THEME-<slug>.
        synth_code = f"DP-THEME-{slugify(name_s)}"
        rows.append({
            "category": f"Themes ({cat_s})",
            "item_code": synth_code,
            "name": name_s,
            "description": f"{cat_s} / {theme_s} theme. Size: {size_s}. Material: {mat_s}.".strip(),
            "fob_usd": 0.0,                # themes are quoted per-project, not catalog-priced
            "remarks": str(remark or "").strip(),
            "source_pricelist": f"{THEME_MANIFEST.name}:row{r_idx}",
            "is_theme": True,
            "theme": theme_s,
            "size_text": size_s,
            "material_text": mat_s,
        })
    log.info("[theme manifest] parsed %d rows", len(rows))
    return rows


# --- Pricing ---------------------------------------------------------------
def price_designpark_row(fob_usd: float, fx: dict, cbm: float = 0.0,
                         kg: float = 0.0) -> dict:
    """USD FOB Busan → landed THB → retail (THB/USD/EUR/SGD).

    When cbm > 0: routes through shipping-automation cost_engine
    origin=japan_korea, method=lcl for a CBM-based Korea LCL estimate.
    When cbm == 0 (no dimension data): falls back to the flat 35% uplift.
    In both cases the Vinci tier-clamp (floor/cap by FOB band) is applied.

    Korea LCL rate: 3,500 THB/CBM (cost_engine japan_korea.lcl).
    Duty: 10% non-China (Korea not covered by ASEAN-China FTA).
    """
    usd_thb = fx.get("USD", 35.0)
    eur_thb = fx.get("EUR", 38.0)
    sgd_thb = fx.get("SGD", 25.0)
    fob_thb = fob_usd * usd_thb
    _g = get_pricing_config() or {}
    th_cust_vat = float(_g.get("th_customer_vat_rate", TH_CUSTOMER_VAT_RATE))
    cfg_gm = float(_g.get("brands", {}).get("designpark", {}).get("gross_margin", GROSS_MARGIN)) \
        if isinstance(_g.get("brands"), dict) else GROSS_MARGIN
    # Pull from flat merged config (get_pricing_config("designpark") merges brand + global)
    from shared.pricing_config import get_pricing_config as _gpc
    dp_cfg = _gpc("designpark") or {}
    gm = float(dp_cfg.get("gross_margin", GROSS_MARGIN))
    duty_rate = float(dp_cfg.get("duty_rate_non_china", DUTY_RATE_NON_CHINA))
    vat_rate = float(dp_cfg.get("thai_vat_rate", THAI_VAT_RATE))

    method_used = "flat_uplift_korea_lcl"
    if cbm and cbm > 0:
        # CBM-based: use shipping-automation cost_engine for Korea LCL
        try:
            est = cost_engine.estimate_landed_cost(
                origin="japan_korea",
                method="lcl",
                goods_value=fob_usd,
                goods_currency="USD",
                cbm=cbm,
                kg=kg,
                duty_rate=duty_rate,
                fx_rates=fx,
            )
            landed_thb_raw = est["total_landed_thb"]
            freight_thb = est["freight"]["thb"]
            duty_thb = est["customs"]["duty_thb"]
            vat_thb = est["customs"]["vat_thb"]
            cif_thb = est["customs"]["cif_thb"]
            method_used = "korea_lcl_cbm"
        except Exception as e:
            log.warning("Korea LCL CBM estimate failed for fob_usd=%.2f cbm=%.3f: %s — "
                        "falling back to flat uplift", fob_usd, cbm, e)
            cbm = 0.0  # fall through to flat path below

    if not (cbm and cbm > 0):
        # Flat-uplift fallback (no CBM data or CBM estimate failed)
        cif_thb = fob_thb * UNMATCHED_LANDED_UPLIFT
        freight_thb = cif_thb - fob_thb
        duty_thb = round(cif_thb * duty_rate, 2)
        vat_thb = round((cif_thb + duty_thb) * vat_rate, 2)
        landed_thb_raw = round(cif_thb + duty_thb + vat_thb, 2)

    # Tier clamp (same floor/cap system as Vinci/Berliner, applied in EUR-band space)
    # DesignPark FOB is in USD — convert to EUR for tier lookup
    eur_fob_equiv = fob_usd * usd_thb / eur_thb  # approx EUR to determine tier
    from shared.landed_pricing import logistics_band, LOGISTICS_TIERS
    tiers_raw = dp_cfg.get("logistics_tiers")
    if tiers_raw:
        tiers = [
            (float("inf") if t.get("fob_eur_max") in (None, "inf") else float(t["fob_eur_max"]),
             float(t["min_pct"]), float(t["max_pct"]))
            for t in tiers_raw
        ]
    else:
        tiers = LOGISTICS_TIERS
    lo_pct, hi_pct = logistics_band(eur_fob_equiv, tiers)
    floor_landed = fob_thb * (1 + lo_pct)
    cap_landed = fob_thb * (1 + hi_pct)
    logistics_clamp = ""
    landed_thb = landed_thb_raw
    if landed_thb < floor_landed:
        landed_thb = floor_landed
        logistics_clamp = "floored"
    elif landed_thb > cap_landed:
        landed_thb = cap_landed
        logistics_clamp = "capped"
    landed_thb = round(landed_thb, 2)

    # Retail: independent per-currency (Task 10). TH customer VAT only on THB.
    retail_thb = round((landed_thb / (1 - gm)) * (1 + th_cust_vat), 2)
    retail_usd = round((landed_thb / usd_thb) / (1 - gm), 2)   # no TH customer VAT
    retail_eur = round((landed_thb / eur_thb) / (1 - gm), 2)
    sg_gst = float(dp_cfg.get("sg_customer_gst_rate", SG_CUSTOMER_GST_RATE))
    sg_reg = bool(dp_cfg.get("sg_nubo_gst_registered", SG_NUBO_GST_REGISTERED))
    sg_gst_mult = (1 + sg_gst) if sg_reg else 1.0
    retail_sgd = round(((landed_thb / sgd_thb) / (1 - gm)) * sg_gst_mult, 2)
    return {
        "fob_usd": round(fob_usd, 2),
        "fob_thb": round(fob_thb, 2),
        "freight_thb": round(freight_thb, 2),
        "duty_thb": round(duty_thb, 2),
        "vat_thb": round(vat_thb, 2),
        "landed_thb": landed_thb,
        "landed_thb_raw": round(landed_thb_raw, 2),
        "logistics_clamp": logistics_clamp,
        "retail_thb": retail_thb,
        "retail_usd": retail_usd,
        "retail_eur": retail_eur,
        "retail_sgd": retail_sgd,
        "gross_margin": gm,
        "th_customer_vat_rate": th_cust_vat,
        "logistics_uplift": UNMATCHED_LANDED_UPLIFT - 1,
        "cbm_used": cbm or 0.0,
        "method": method_used,
        "formula_version": FORMULA_VERSION,
    }


# --- Firestore write -------------------------------------------------------
def build_product_doc(row: dict, fx: dict) -> dict:
    pricing = price_designpark_row(row.get("fob_usd") or 0.0, fx) if row.get("fob_usd") else {
        "fob_usd": 0.0,
        "formula_version": FORMULA_VERSION,
        "method": "no_price_in_pricelist",
    }
    handle = to_handle(row["item_code"], row["name"])
    doc: dict = {
        "handle": handle,
        "slug": SLUG,
        "name": row["name"],
        "item_code": row["item_code"],
        "category": row["category"],
        "description": row.get("description") or "",
        "remarks": row.get("remarks") or "",
        "source_pricelist": row["source_pricelist"],
        "pricing": pricing,
        # Status set provisionally; shape_designpark_to_medusa_schema.py will
        # promote to "active" once at least one image is joined in.
        "status": "draft_no_images",
    }
    if row.get("is_theme"):
        doc["is_theme"] = True
        doc["theme"] = row.get("theme") or ""
        doc["size_text"] = row.get("size_text") or ""
        doc["material_text"] = row.get("material_text") or ""
    return doc


def write_firestore(docs: list[dict], dry: bool) -> tuple[int, int]:
    if dry:
        log.info("[DRY] would write %d docs to vendors/%s/products in db=%s",
                 len(docs), SLUG, VENDORS_DB)
        for d in docs[:5]:
            log.info("    %s | %s | fob_usd=%s | retail_usd=%s",
                     d["handle"], d["name"][:40],
                     d["pricing"].get("fob_usd"), d["pricing"].get("retail_usd"))
        if len(docs) > 5:
            log.info("    ... (%d more)", len(docs) - 5)
        return (len(docs), 0)

    db = firestore.Client(project=PROJECT, database=VENDORS_DB)
    coll = db.collection("vendors").document(SLUG).collection("products")
    n_write, n_skip = 0, 0
    # Batch in groups of 400 (Firestore limit 500).
    batch = db.batch()
    pending = 0
    for d in docs:
        ref = coll.document(d["handle"])
        batch.set(ref, d, merge=True)
        pending += 1
        n_write += 1
        if pending >= 400:
            batch.commit()
            batch = db.batch()
            pending = 0
    if pending:
        batch.commit()
    return (n_write, n_skip)


# --- Main ------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser()
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--dry-run", action="store_true")
    g.add_argument("--apply", action="store_true")
    ap.add_argument("--limit", type=int, default=None, help="cap rows for smoke tests")
    ap.add_argument("--skip-themes", action="store_true")
    ap.add_argument("--skip-components", action="store_true")
    ap.add_argument("--dump-csv", type=str, default=None, help="write priced rows to a CSV")
    args = ap.parse_args()
    dry = args.dry_run

    if not PRICELIST_USD.exists() and not args.skip_components:
        log.error("pricelist not found: %s", PRICELIST_USD)
        return 2

    # FX snapshot (live; falls back internally to static rates).
    fx = get_fx_rates()
    log.info("FX snapshot: USD=%.4f THB/USD, EUR=%.4f THB/EUR", fx.get("USD", 0), fx.get("EUR", 0))

    # 1) Components.
    rows: list[dict] = []
    if not args.skip_components:
        wb = openpyxl.load_workbook(PRICELIST_USD, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows.extend(parse_sheet(sn, ws))

    # 2) Themes.
    if not args.skip_themes:
        rows.extend(parse_theme_manifest())

    if args.limit:
        rows = rows[: args.limit]
        log.info("limited to %d rows", len(rows))

    # 3) Build docs.
    docs = [build_product_doc(r, fx) for r in rows]
    log.info("built %d product docs", len(docs))

    # 4) Optional CSV dump (audit trail).
    if args.dump_csv:
        out = Path(args.dump_csv)
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh)
            w.writerow([
                "handle", "item_code", "category", "name",
                "fob_usd", "landed_thb", "retail_thb", "retail_usd", "retail_eur",
                "source_pricelist",
            ])
            for d in docs:
                p = d.get("pricing", {})
                w.writerow([
                    d["handle"], d["item_code"], d["category"], d["name"],
                    p.get("fob_usd"), p.get("landed_thb"),
                    p.get("retail_thb"), p.get("retail_usd"), p.get("retail_eur"),
                    d["source_pricelist"],
                ])
        log.info("wrote audit CSV: %s", out)

    # 5) Write Firestore.
    n_write, n_skip = write_firestore(docs, dry)
    log.info("%s %d docs (skipped %d) → vendors/%s/products",
             "[DRY]" if dry else "wrote", n_write, n_skip, SLUG)
    return 0


if __name__ == "__main__":
    sys.exit(main())
