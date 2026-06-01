"""
extract_po_images.py — Pull the embedded product photos out of a Wisdom PO Excel
and map each one to its item code via the drawing anchors.

The 2026-06-01 Dulwich PO embeds one product photo per line item in column C.
They are stored as EMF/WMF/PNG inside the .xlsx (xl/media/*). openpyxl can't
read EMF, so we go straight to the zip: drawing1.xml anchors each image to a
`<xdr:from><xdr:row>` cell, and drawing1.xml.rels maps the embed rId to the
media file. We pair that row with the PO line-item code (same row order as the
sheet) to name each extracted file `<item_code>.<ext>`.

EMF files still need rasterizing — run convert_po_emf.ps1 (Windows GDI+) next.

Usage:
    python wisdom-catalog/extract_po_images.py
    python wisdom-catalog/extract_po_images.py --po "C:/path/PO.xlsx" --out wisdom-catalog/exports/po_images_raw
"""
from __future__ import annotations

import argparse
import json
import os
import re
import zipfile

DEFAULT_PO = (
    r"C:\Users\Eukrit\My Drive\Partners Playground\Wisdom Playground"
    r"\2026-06-01 PO 20260601 Dulwich Singapore.xlsx"
)
DEFAULT_OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports", "po_images_raw")
DATA_START_ROW = 11  # 0-based; PO line items begin on sheet row 12 (1-based)


def po_codes(path: str) -> dict[int, str]:
    """row (0-based) -> item_code for each PO line item."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    out: dict[int, str] = {}
    i = DATA_START_ROW
    for r in rows[DATA_START_ROW:]:
        code = r[0] if r else None
        if not code or not str(code).strip():
            i += 1
            continue
        code = str(code).strip()
        if code.lower() == "total":
            break
        out[i] = code
        i += 1
    return out


def anchor_map(z: zipfile.ZipFile) -> list[tuple[int, str]]:
    """[(from_row_0based, media_path)] for each anchored image, sorted by row."""
    draw = z.read("xl/drawings/drawing1.xml").decode("utf-8", "replace")
    rels = z.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8", "replace")
    rid2media = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
    anchors = re.findall(r"<xdr:(?:two|one)CellAnchor.*?</xdr:(?:two|one)CellAnchor>", draw, re.S)
    out = []
    for a in anchors:
        mr = re.search(r"<xdr:from>.*?<xdr:row>(\d+)</xdr:row>", a, re.S)
        me = re.search(r'r:embed="([^"]+)"', a)
        if mr and me and me.group(1) in rid2media:
            out.append((int(mr.group(1)), rid2media[me.group(1)].replace("../", "xl/")))
    out.sort()
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--po", default=DEFAULT_PO)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)
    row2code = po_codes(args.po)
    z = zipfile.ZipFile(args.po)
    pairs = anchor_map(z)

    manifest = []
    for from_row, media in pairs:
        code = row2code.get(from_row)
        data = z.read(media)
        ext = media.rsplit(".", 1)[-1].lower()
        fn = f"{code}.{ext}" if code else f"row{from_row + 1}.{ext}"
        with open(os.path.join(args.out, fn), "wb") as fh:
            fh.write(data)
        manifest.append({
            "row_1based": from_row + 1, "code": code, "media": media,
            "ext": ext, "bytes": len(data), "file": fn,
        })

    with open(os.path.join(args.out, "manifest.json"), "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, indent=2, ensure_ascii=False)

    exts: dict[str, int] = {}
    for m in manifest:
        exts[m["ext"]] = exts.get(m["ext"], 0) + 1
    print(f"Extracted {len(manifest)} images -> {args.out}")
    print(f"  ext breakdown: {exts}")
    print(f"  unmapped rows: {[m['row_1based'] for m in manifest if not m['code']]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
