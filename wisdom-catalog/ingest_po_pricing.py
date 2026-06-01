"""
ingest_po_pricing.py — Ingest a Wisdom Proforma-Invoice / PO price list into the
catalog (Firestore products_wisdom + Medusa) and record the PO as a vendor quotation.

Unlike update_pricing.py (which recomputes retail from the FOB already stored on
each product), this script *sets* pricing.fob_usd from a vendor PO Excel — the PO
is the authoritative source of the negotiated Ex-work / FOB price. It then reuses
the canonical pricing math (shared/wisdom_pricing.py) to recompute landed + retail
(THB/USD/SGD) and pushes the result to Medusa, exactly like update_pricing.py.

Default PO: the 2026-06-01 Dulwich Singapore proforma (PO 2026060101) from TUMACO
LIMITED. The sheet columns are:
    A Product Code | B Description | C Image | D Material Spec |
    E Ex-work Shanghai (USD) Per Unit | F QTY | G Unit Volume (CBM) |
    H Volume (CBM) | I Amount(USD) | J Remarks
Header is on row 10 (0-based 10); data rows start at row 11; stop at the "Total" row.

Writes (only with --write):
  * products_wisdom/{item_code}.pricing.*  (fob_usd + recomputed landed/retail + rates)
  * products_wisdom/{item_code}.volume_cbm (from the PO unit volume, when present)
  * leka_vendor_quotations/{quotation_doc_id}  (the PO snapshot)
  * Medusa variant prices (THB retail + USD FOB)  — skip with --skip-medusa
  * wisdom-catalog/exports/<po>-priced.json  (handoff for the leka-projects R2 session)

Usage:
    python wisdom-catalog/ingest_po_pricing.py --dry-run
    python wisdom-catalog/ingest_po_pricing.py --write
    python wisdom-catalog/ingest_po_pricing.py --write --skip-medusa
    python wisdom-catalog/ingest_po_pricing.py --po "C:/path/to/other-po.xlsx" \
        --po-number 2026060101 --po-date 2026-06-01 --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from shared.wisdom_pricing import (  # noqa: E402
    compute_wisdom_retail,
    pricing_metadata,
    get_usd_thb,
    get_sgd_thb,
)

GCP_PROJECT = "ai-agents-go"
CATALOG_DB = "leka-product-catalogs"
COLLECTION = "products_wisdom"
QUOTATION_COLLECTION = "leka_vendor_quotations"
BATCH_SIZE = 400

# Default PO — 2026-06-01 Dulwich Singapore proforma invoice.
DEFAULT_PO = (
    r"C:\Users\Eukrit\My Drive\Partners Playground\Wisdom Playground"
    r"\2026-06-01 PO 20260601 Dulwich Singapore.xlsx"
)
DEFAULT_PO_NUMBER = "2026060101"
DEFAULT_PO_DATE = "2026-06-01"
DEFAULT_VENDOR = "TUMACO LIMITED"
DEFAULT_SOURCE = "po_dulwich_singapore_2026-06-01"
DEFAULT_PROJECT = "dulwich-singapore"
PRICE_TERM = "Ex-work Shanghai, China"

EXPORTS_DIR = REPO_ROOT / "wisdom-catalog" / "exports"

# Column indices (0-based) on the PO sheet.
COL_CODE = 0
COL_DESC = 1
COL_FOB = 4
COL_QTY = 5
COL_UNIT_CBM = 6
COL_AMOUNT = 8
COL_REMARKS = 9
DATA_START_ROW = 11  # 0-based; row 10 is the header


def _credentials_path() -> str | None:
    for p in (
        r"C:\Users\Eukrit\OneDrive\Documents\Claude Code\Credentials Claude Code"
        r"\ai-agents-go-claude-sa.json",
        r"C:\Users\eukri\OneDrive\Documents\Claude Code\Credentials Claude Code"
        r"\ai-agents-go-claude-sa.json",
    ):
        if os.path.exists(p):
            return p
    return None


def _firestore_client():
    from google.cloud import firestore
    if not os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"):
        cred = _credentials_path()
        if cred:
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = cred
    return firestore.Client(project=GCP_PROJECT, database=CATALOG_DB)


def _norm(s) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper()) if s else ""


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_po(path: Path) -> list[dict]:
    """Parse the PO Excel into a list of line-item dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    items: list[dict] = []
    for r in rows[DATA_START_ROW:]:
        if not r:
            continue
        code = r[COL_CODE]
        if not code or not str(code).strip():
            continue
        code = str(code).strip()
        if code.lower() == "total":
            break
        fob = _num(r[COL_FOB] if len(r) > COL_FOB else None)
        if fob is None or fob <= 0:
            continue
        items.append({
            "item_code": code,
            "description": (str(r[COL_DESC]).strip() if len(r) > COL_DESC and r[COL_DESC] else None),
            "fob_usd": round(fob, 2),
            "qty": _num(r[COL_QTY] if len(r) > COL_QTY else None),
            "volume_cbm": _num(r[COL_UNIT_CBM] if len(r) > COL_UNIT_CBM else None),
            "amount_usd": _num(r[COL_AMOUNT] if len(r) > COL_AMOUNT else None),
            "remarks": (str(r[COL_REMARKS]).strip() if len(r) > COL_REMARKS and r[COL_REMARKS] else None),
        })
    return items


def match_products(db, items: list[dict]) -> tuple[list[dict], list[dict]]:
    """Resolve each PO item to a products_wisdom doc id (exact, then normalized)."""
    docs: dict[str, dict] = {}
    for d in db.collection(COLLECTION).stream():
        docs[d.id] = d.to_dict() or {}
    norm_index = {_norm(i): i for i in docs}

    matched, missing = [], []
    for it in items:
        code = it["item_code"]
        doc_id = code if code in docs else norm_index.get(_norm(code))
        if doc_id:
            cur = (docs[doc_id].get("pricing") or {}).get("fob_usd")
            matched.append({**it, "doc_id": doc_id, "current_fob": cur})
        else:
            missing.append(it)
    return matched, missing


def compute_rows(matched: list[dict], usd_thb: float, sgd_thb: float,
                 price_date: str) -> list[dict]:
    """Recompute landed/retail for each matched item and build update rows."""
    rows = []
    for m in matched:
        row = compute_wisdom_retail(m["fob_usd"], usd_thb, sgd_thb)
        if not row:
            continue
        row.item_code = m["doc_id"]
        rows.append({
            "item_code": m["doc_id"],          # Medusa lookup key (legacy_sku)
            "po_code": m["item_code"],
            "fob_usd": m["fob_usd"],
            "current_fob": m.get("current_fob"),
            "volume_cbm": m.get("volume_cbm"),
            "qty": m.get("qty"),
            "pricing_update": pricing_metadata(row, price_date),
            "retail_sgd": row.retail_sgd,
        })
    return rows


def write_firestore(db, rows: list[dict], price_date: str, dry_run: bool) -> int:
    batch = db.batch()
    count = 0
    for r in rows:
        ref = db.collection(COLLECTION).document(r["item_code"])
        p = r["pricing_update"]
        update: dict = {
            "pricing.fob_usd":          p["fob_usd"],
            "pricing.currency":         p.get("currency", "USD"),
            "pricing.landed_thb":       p["landed_thb"],
            "pricing.retail_thb":       p["retail_thb"],
            "pricing.retail_usd":       p["retail_usd"],
            "pricing.retail_sgd":       p["retail_sgd"],
            "pricing.duty_thb":         p["duty_thb"],
            "pricing.vat_thb":          p["vat_thb"],
            "pricing.usd_thb":          p["usd_thb"],
            "pricing.sgd_thb":          p["sgd_thb"],
            "pricing.import_duty_rate": p["import_duty_rate"],
            "pricing.thai_vat_rate":    p["thai_vat_rate"],
            "pricing.gross_margin":     p["gross_margin"],
            "pricing.price_date":       price_date,
            "pricing.price_source":     DEFAULT_SOURCE,
            "updated_at":               datetime.now(timezone.utc),
        }
        if r.get("volume_cbm"):
            update["volume_cbm"] = r["volume_cbm"]
        if dry_run:
            count += 1
            continue
        batch.update(ref, update)
        count += 1
        if count % BATCH_SIZE == 0:
            batch.commit()
            batch = db.batch()
    if not dry_run and count % BATCH_SIZE != 0:
        batch.commit()
    return count


def write_quotation(db, items: list[dict], args, dry_run: bool) -> str:
    """Record the PO as a single leka_vendor_quotations document."""
    doc_id = f"wisdom-PO-{args.po_number}"
    total_usd = round(sum((it.get("amount_usd") or 0) for it in items), 2)
    payload = {
        "quotation_id":  args.po_number,
        "po_number":     args.po_number,
        "brand":         "wisdom",
        "vendor_name":   args.vendor,
        "date":          args.po_date,
        "source":        args.source,
        "project":       args.project,
        "price_term":    PRICE_TERM,
        "currency":      "USD",
        "total_usd":     total_usd,
        "items": [
            {
                "item_code":  it["item_code"],
                "fob_usd":    it["fob_usd"],
                "volume_cbm": it.get("volume_cbm"),
                "qty":        it.get("qty"),
                "amount_usd": it.get("amount_usd"),
                "remarks":    it.get("remarks"),
            }
            for it in items
        ],
        "created_at": datetime.now(timezone.utc),
    }
    if not dry_run:
        db.collection(QUOTATION_COLLECTION).document(doc_id).set(payload)
    return doc_id


def write_handoff(rows: list[dict], args, dry_run: bool) -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out = EXPORTS_DIR / f"dulwich-po-{args.po_number}-priced.json"
    payload = {
        "po_number": args.po_number,
        "po_date": args.po_date,
        "vendor": args.vendor,
        "project": args.project,
        "brand": "wisdom",
        "wisdom_fob_to_sgd": round(104.09 * 1.05 / 24.6, 4),  # mirrors build_r2_draft_order
        "items": [
            {
                "item_code": r["item_code"],
                "fob_usd": r["fob_usd"],
                "retail_sgd": r["retail_sgd"],
                "qty": r.get("qty"),
                "volume_cbm": r.get("volume_cbm"),
                "was_unpriced": r.get("current_fob") in (None, 0),
            }
            for r in rows
        ],
    }
    if not dry_run:
        out.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Ingest a Wisdom PO price list")
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--dry-run", action="store_true", help="Preview; no writes")
    g.add_argument("--write", action="store_true", help="Commit Firestore + quotation + Medusa")
    ap.add_argument("--skip-medusa", action="store_true", help="Firestore/quotation only")
    ap.add_argument("--po", default=DEFAULT_PO, help="Path to the PO Excel")
    ap.add_argument("--po-number", default=DEFAULT_PO_NUMBER)
    ap.add_argument("--po-date", default=DEFAULT_PO_DATE, help="YYYY-MM-DD")
    ap.add_argument("--vendor", default=DEFAULT_VENDOR)
    ap.add_argument("--source", default=DEFAULT_SOURCE)
    ap.add_argument("--project", default=DEFAULT_PROJECT)
    ap.add_argument("--usd-thb", type=float, default=None)
    ap.add_argument("--sgd-thb", type=float, default=None)
    ap.add_argument("--medusa-url",
                    default=os.environ.get("MEDUSA_BACKEND_URL",
                                           "https://leka-medusa-backend-538978391890.asia-southeast1.run.app"))
    args = ap.parse_args()
    dry = args.dry_run

    usd_thb = args.usd_thb or get_usd_thb()
    sgd_thb = args.sgd_thb or get_sgd_thb()
    price_date = args.po_date

    print(f"Wisdom PO ingest — {args.po_number} ({args.po_date}), vendor={args.vendor}")
    print(f"  PO file: {args.po}")
    print(f"  Rates:   USD/THB {usd_thb:.4f}   SGD/THB {sgd_thb:.4f}")
    print(f"  Mode:    {'DRY-RUN' if dry else 'WRITE'}{'  (skip-medusa)' if args.skip_medusa else ''}\n")

    items = read_po(Path(args.po))
    print(f"Parsed {len(items)} PO line items "
          f"(total USD {sum((it.get('amount_usd') or 0) for it in items):,.2f})\n")
    if not items:
        print("No line items parsed — aborting.")
        return 1

    db = _firestore_client()
    print(f"Loading {COLLECTION}…")
    matched, missing = match_products(db, items)
    print(f"  Matched: {len(matched)}   Missing: {len(missing)}")
    if missing:
        for it in missing:
            print(f"    MISSING  {it['item_code']}  (FOB {it['fob_usd']})")
    print()

    rows = compute_rows(matched, usd_thb, sgd_thb, price_date)

    print("Per-code pricing:")
    newly = 0
    for r in rows:
        p = r["pricing_update"]
        flag = ""
        if r.get("current_fob") in (None, 0):
            flag = "  [NEW FOB]"
            newly += 1
        elif r.get("current_fob") is not None and abs(r["current_fob"] - r["fob_usd"]) >= 0.01:
            flag = f"  [Δ was {r['current_fob']:.2f}]"
        print(f"  {r['po_code']:18s} FOB ${r['fob_usd']:>8.2f}  "
              f"landed ฿{p['landed_thb']:>9,.0f}  "
              f"retail ฿{p['retail_thb']:>9,.0f} / S${r['retail_sgd']:>8,.2f}{flag}")
    print(f"\n  {len(rows)} priced  |  {newly} newly priced (had no FOB)\n")

    # --- Firestore products ---
    n = write_firestore(db, rows, price_date, dry_run=dry)
    print(f"{'[dry] would update' if dry else 'Updated'} {n} products_wisdom docs")

    # --- Quotation record ---
    qid = write_quotation(db, items, args, dry_run=dry)
    print(f"{'[dry] would record' if dry else 'Recorded'} quotation {QUOTATION_COLLECTION}/{qid}")

    # --- Handoff JSON ---
    out = write_handoff(rows, args, dry_run=dry)
    print(f"{'[dry] would write' if dry else 'Wrote'} handoff {out}")

    # --- Medusa push (reuse update_pricing.update_medusa) ---
    if not args.skip_medusa:
        # Reuse update_pricing.update_medusa (sibling module in wisdom-catalog/).
        sys.path.insert(0, str(REPO_ROOT / "wisdom-catalog"))
        import update_pricing as upmod  # noqa: E402
        from shared.medusa_importer import MedusaImporter
        if dry:
            upmod.update_medusa(None, rows, dry_run=True)
        else:
            os.environ.setdefault("MEDUSA_BACKEND_URL", args.medusa_url)
            client = MedusaImporter(base_url=args.medusa_url)
            updated, skipped = upmod.update_medusa(client, rows, dry_run=False)
            print(f"  Medusa: {updated} variants updated, {skipped} skipped/not-found")
    else:
        print("  (Medusa push skipped)")

    print(f"\nDone. PO {args.po_number}: {len(rows)} Wisdom products priced "
          f"({newly} newly).{'  [DRY-RUN — nothing written]' if dry else ''}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
