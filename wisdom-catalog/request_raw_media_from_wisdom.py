"""Export the full Wisdom catalog (5,071 SKUs from Firestore products_wisdom)
to an Excel file, then send a vendor-request email to the Wisdom team via
Gmail DWD impersonation as eukrit@goco.bz.

Reads products_wisdom from the `leka-product-catalogs` Firestore database in
GCP project `ai-agents-go`. Output xlsx is timestamped under
wisdom-catalog/exports/.

Usage:
    python wisdom-catalog/request_raw_media_from_wisdom.py            # export + send
    python wisdom-catalog/request_raw_media_from_wisdom.py --no-send  # export only
    python wisdom-catalog/request_raw_media_from_wisdom.py --dry-run  # build draft, don't send

Credentials: GOOGLE_APPLICATION_CREDENTIALS must point at the
ai-agents-go SA JSON for DWD impersonation. ADC is used for Firestore.
"""
from __future__ import annotations

import argparse
import base64
import datetime as dt
import logging
import os
import sys
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
log = logging.getLogger("wisdom_request")

GCP_PROJECT = "ai-agents-go"
FIRESTORE_DATABASE = "leka-product-catalogs"
PRODUCTS_COLLECTION = "products_wisdom"

GMAIL_FROM = "Eukrit Kongtaveelert <eukrit@goco.bz>"
GMAIL_SUBJECT_USER = "eukrit@goco.bz"
GMAIL_TO = [
    "alex@huasenwei.com",
    "amanda@huasenwei.com",
    "martin_zhu@huasenwei.com",
    "martin@huasenwei.com",
]

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(THIS_DIR, "exports")


def fetch_all_products() -> list[dict]:
    """Pull every doc in products_wisdom."""
    from google.cloud import firestore
    db = firestore.Client(project=GCP_PROJECT, database=FIRESTORE_DATABASE)
    log.info("Querying %s/%s …", FIRESTORE_DATABASE, PRODUCTS_COLLECTION)
    out = []
    for snap in db.collection(PRODUCTS_COLLECTION).stream():
        d = snap.to_dict() or {}
        d["_doc_id"] = snap.id
        out.append(d)
    log.info("  %d products loaded.", len(out))
    return out


COLUMNS = [
    ("item_code",        "Item Code",            16, "left"),
    ("description",      "Description (EN)",     46, "left"),
    ("description_cn",   "Description (CN / 中文描述)", 38, "left"),
    ("category",         "Category",             16, "left"),
    ("subcategory",      "Subcategory",          18, "left"),
    ("material",         "Material",             24, "left"),
    ("dimensions",       "Dimensions (raw)",     28, "left"),
    ("weight_kg",        "Weight (kg)",          12, "right"),
    ("volume_cbm",       "Volume (m³)",          12, "right"),
    ("fob_usd",          "Current FOB (USD)",    16, "right"),
    ("currency",         "Currency",             10, "left"),
    ("price_date",       "Price date",           14, "left"),
    ("image_count",      "Images on file",       14, "right"),
    ("catalog_source",   "Catalog source",       28, "left"),
    ("catalog_page",     "Catalog page",         14, "right"),
    ("needs",            "Needs (please fill)",  44, "left"),
]


def _flatten(p: dict) -> dict:
    """Pull the fields out of nested objects."""
    pricing = p.get("pricing") or {}
    dims = p.get("dimensions") or {}
    images = p.get("images") or []
    image_count = len([i for i in images if isinstance(i, dict) and i.get("url")])
    needs = []
    if not p.get("description"):
        needs.append("EN description")
    if not pricing.get("fob_usd"):
        needs.append("FOB price")
    if not dims.get("raw") and not p.get("dimensions_raw"):
        needs.append("dimensions")
    if not p.get("weight_kg") and not p.get("weight"):
        needs.append("weight")
    if image_count == 0:
        needs.append("image")
    return {
        "item_code": p.get("_doc_id") or p.get("item_code") or "",
        "description": p.get("description") or "",
        "description_cn": p.get("description_cn") or "",
        "category": p.get("category") or "",
        "subcategory": p.get("subcategory") or "",
        "material": p.get("material") or "",
        "dimensions": dims.get("raw") or p.get("dimensions_raw") or "",
        "weight_kg": p.get("weight_kg") or p.get("weight"),
        "volume_cbm": p.get("volume_cbm") or p.get("volume"),
        "fob_usd": pricing.get("fob_usd"),
        "currency": pricing.get("currency") or "",
        "price_date": pricing.get("price_date") or "",
        "image_count": image_count,
        "catalog_source": p.get("catalog_source") or "",
        "catalog_page": p.get("catalog_page"),
        "needs": ", ".join(needs) if needs else "",
    }


def build_xlsx(products: list[dict], out_path: str) -> dict:
    log.info("Building Excel …")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wisdom SKUs"

    header_fill = PatternFill("solid", fgColor="182557")
    header_font = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
    body_font = Font(name="Calibri", size=10)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="top")

    # Header
    for col_idx, (key, header, width, align) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_left
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Sort: item_code ascending
    rows = sorted([_flatten(p) for p in products],
                  key=lambda r: r["item_code"] or "")

    counts = {
        "total": len(rows),
        "has_description": 0,
        "has_fob_price": 0,
        "has_dimensions": 0,
        "has_weight": 0,
        "has_image": 0,
        "has_all_needs": 0,
        "missing_any": 0,
    }

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _h, _w, align) in enumerate(COLUMNS, start=1):
            value = row.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = align_right if align == "right" else align_left
        if row.get("description"):
            counts["has_description"] += 1
        if row.get("fob_usd"):
            counts["has_fob_price"] += 1
        if row.get("dimensions"):
            counts["has_dimensions"] += 1
        if row.get("weight_kg"):
            counts["has_weight"] += 1
        if row.get("image_count"):
            counts["has_image"] += 1
        if not row.get("needs"):
            counts["has_all_needs"] += 1
        else:
            counts["missing_any"] += 1

    # Sheet 2 — coverage summary
    sheet2 = wb.create_sheet("Coverage summary")
    sheet2["A1"] = "Field"
    sheet2["B1"] = "Have"
    sheet2["C1"] = "Missing"
    sheet2["D1"] = "Coverage %"
    for c in ("A1", "B1", "C1", "D1"):
        sheet2[c].fill = header_fill
        sheet2[c].font = header_font
    rows_summary = [
        ("EN description", counts["has_description"]),
        ("FOB USD price", counts["has_fob_price"]),
        ("Dimensions", counts["has_dimensions"]),
        ("Weight (kg)", counts["has_weight"]),
        ("Product image", counts["has_image"]),
        ("All fields complete", counts["has_all_needs"]),
    ]
    total = counts["total"]
    for i, (label, have) in enumerate(rows_summary, start=2):
        sheet2.cell(row=i, column=1, value=label)
        sheet2.cell(row=i, column=2, value=have)
        sheet2.cell(row=i, column=3, value=total - have)
        sheet2.cell(row=i, column=4, value=f"{100 * have / max(1, total):.1f}%")
    sheet2.column_dimensions["A"].width = 30
    sheet2.column_dimensions["B"].width = 10
    sheet2.column_dimensions["C"].width = 10
    sheet2.column_dimensions["D"].width = 14
    sheet2.cell(row=i + 2, column=1, value=f"Total SKUs: {total}")
    sheet2.cell(row=i + 2, column=1).font = Font(bold=True)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    log.info("Saved %s", out_path)
    return counts


def build_email_body(counts: dict) -> str:
    total = counts["total"]
    return f"""Dear Alex, Amanda, Martin Zhu, and Martin,

We are preparing the next refresh of our Leka Project storefront, which carries Wisdom's outdoor-play, furniture, and playground range. To bring our records fully up to date, we would like to request the latest raw media for every Wisdom item code in our database.

Attached is an Excel file (wisdom_raw_media_request_{dt.date.today().isoformat()}.xlsx) listing all {total} Wisdom product codes we currently track, split across two sheets:

  1. "Wisdom SKUs" — one row per item code, with the data we currently have on file. The "Needs (please fill)" column highlights what is missing for that SKU.
  2. "Coverage summary" — quick rollup of where we have gaps.

For each item code we would like:

  • Current FOB USD price (and validity period if applicable).
  • Latest product dimensions (and package dimensions / carton size if available).
  • Product weight (kg) and packed / carton gross weight (kg) — for every SKU.
  • One or more high-resolution single-product images — clean photos of the product on its own, ideally not catalog-page layouts where many products share one frame.

Where the data on file is still correct, please just leave the cell as-is. Where we have gaps or your information has changed, please update the cell directly in the file and return it to us.

For the product photos, we would like to request a single shared Google Drive folder (or Dropbox / WeTransfer) containing high-resolution images for all {total}+ Wisdom SKUs, organized by item code (one sub-folder or file name per code). This is important to us: most of our current images came from shared catalog-page layouts where a single photo covers many products, so we cannot reliably assign the correct image to each individual code. A shared drive of per-code photos lets us bulk-download everything and automatically map each picture to its SKU. Please make sure the folder covers the full range of 5,000+ item codes, not just a selection.

Coverage right now (based on what we have on file):

  • EN description: {counts['has_description']} / {total}
  • FOB USD price: {counts['has_fob_price']} / {total}
  • Dimensions: {counts['has_dimensions']} / {total}
  • Weight (kg): {counts['has_weight']} / {total}
  • Product image: {counts['has_image']} / {total}

Please let us know if you would prefer to send the data back in a different format (Google Sheet, separate price list, etc.) — we can accommodate whatever is easiest for your team.

Thank you very much for your support.

Best regards,
Eukrit Kongtaveelert
Director, GO Corporation Co., Ltd.
eukrit@goco.bz
"""


def send_email(xlsx_path: str, body: str, *, dry_run: bool) -> dict:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    sa_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not sa_path or not os.path.isfile(sa_path):
        raise SystemExit(
            "GOOGLE_APPLICATION_CREDENTIALS env var must point at the "
            "ai-agents-go SA JSON for DWD impersonation."
        )

    creds = service_account.Credentials.from_service_account_file(
        sa_path,
        scopes=["https://www.googleapis.com/auth/gmail.send"],
        subject=GMAIL_SUBJECT_USER,
    )
    service = build("gmail", "v1", credentials=creds, cache_discovery=False)

    msg = MIMEMultipart()
    msg["From"] = GMAIL_FROM
    msg["To"] = ", ".join(GMAIL_TO)
    msg["Subject"] = (
        f"Wisdom — request for updated prices, dimensions and product images "
        f"({dt.date.today().isoformat()})"
    )
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(xlsx_path, "rb") as fh:
        part = MIMEApplication(
            fh.read(),
            _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        part.add_header("Content-Disposition",
                        "attachment",
                        filename=os.path.basename(xlsx_path))
        msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()

    if dry_run:
        log.info("[dry-run] Would send to: %s", msg["To"])
        log.info("[dry-run] Subject: %s", msg["Subject"])
        log.info("[dry-run] Attachment: %s (%.1f KB)",
                 os.path.basename(xlsx_path),
                 os.path.getsize(xlsx_path) / 1024)
        return {"dry_run": True}

    log.info("Sending email …")
    sent = service.users().messages().send(userId="me", body={"raw": raw}).execute()
    log.info("Sent — message id: %s", sent.get("id"))
    return sent


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--no-send", action="store_true",
                        help="Export the Excel only; don't email.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Build the draft (incl. attachment) but don't send.")
    parser.add_argument("--xlsx",
                        default=os.path.join(EXPORT_DIR,
                                             f"wisdom_raw_media_request_{dt.date.today().isoformat()}.xlsx"))
    args = parser.parse_args()

    products = fetch_all_products()
    counts = build_xlsx(products, args.xlsx)

    log.info("Excel coverage rollup:")
    for k, v in counts.items():
        log.info("  %-22s %s", k, v)

    if args.no_send:
        log.info("--no-send → stopping after export.")
        return

    body = build_email_body(counts)
    send_email(args.xlsx, body, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
